"""Excel export router — tüm veriyi multi-sheet XLSX olarak dışa aktarır."""
import io
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import db
from dependencies import get_current_user, write_audit

router = APIRouter(prefix="/export", tags=["export"])


def _header_style(ws, row=1):
    """Apple/Notion stili minimal header."""
    for cell in ws[row]:
        cell.font = Font(bold=True, color="86868B", size=10, name="Helvetica")
        cell.fill = PatternFill("solid", fgColor="FAFAFA")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="E5E5EA"))


def _auto_width(ws, max_w=50):
    """Sütun genişliklerini içeriğe göre otomatik ayarla."""
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 10
        for c in col_cells[:200]:
            try:
                if c.value is not None:
                    max_len = max(max_len, min(len(str(c.value)) + 2, max_w))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len


def _write_sheet(wb, name, rows, columns):
    """rows: list[dict], columns: list[(field, label)]"""
    ws = wb.create_sheet(name[:31])  # Excel sheet name max 31 char
    headers = [label for _, label in columns]
    ws.append(headers)
    _header_style(ws, 1)
    for r in rows:
        ws.append([r.get(field) for field, _ in columns])
    _auto_width(ws)
    return ws


@router.get("/excel")
async def export_excel(user: dict = Depends(get_current_user)):
    """Tüm veriyi Excel olarak indirir — 1 dosyada tüm sheet'ler."""
    wb = Workbook()
    wb.remove(wb.active)  # default sheet'i sil

    # 1. Borçlar
    payables = await db.payables.find({"kind": "PAYABLE"}, {"_id": 0}).sort("due_date", -1).to_list(10000)
    _write_sheet(wb, "Borçlar", payables, [
        ("due_date", "Vade"), ("vendor", "Tedarikçi"), ("ship", "Gemi/Birim"),
        ("expense_type", "Masraf Türü"), ("expense_code", "Masraf Kodu"),
        ("description", "Açıklama"), ("original_amount", "Tutar"),
        ("currency", "Döviz"), ("usd_amount", "USD Karşılığı"),
        ("status", "Durum"), ("armator", "Armatör"),
        ("country", "Ülke"), ("year", "Yıl"), ("month", "Ay"),
        ("created_at", "Eklenme Tarihi"), ("created_by", "Ekleyen"),
    ])

    # 2. Alacaklar
    receivables = await db.payables.find({"kind": "RECEIVABLE"}, {"_id": 0}).sort("due_date", -1).to_list(10000)
    _write_sheet(wb, "Alacaklar", receivables, [
        ("due_date", "Vade"), ("vendor", "Müşteri"), ("ship", "Gemi/Birim"),
        ("description", "Açıklama"), ("original_amount", "Tutar"),
        ("currency", "Döviz"), ("usd_amount", "USD Karşılığı"),
        ("status", "Durum"), ("created_at", "Eklenme"),
    ])

    # 3. Ödemeler/Tahsilatlar
    payments = await db.payments.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
    _write_sheet(wb, "Ödemeler", payments, [
        ("date", "Tarih"), ("type", "Tip"), ("vendor", "Firma"),
        ("description", "Açıklama"), ("paying_company", "Ödeyen Şirket"),
        ("payment_method", "Banka/Kasa"), ("ship", "Gemi/Birim"),
        ("amount", "Tutar"), ("currency", "Döviz"), ("fx_rate", "Kur"),
        ("usd_amount", "USD Karşılığı"), ("approved", "Onaylı"),
        ("created_at", "Eklenme"), ("created_by", "Ekleyen"),
    ])

    # 4. Cari Hesaplar Özet
    payables_agg = await db.payables.aggregate([
        {"$match": {"kind": "PAYABLE"}},
        {"$group": {"_id": "$vendor", "debt": {"$sum": "$usd_amount"}, "count": {"$sum": 1}}}
    ]).to_list(5000)
    payments_agg = await db.payments.aggregate([
        {"$match": {"type": "TEDİYE"}},
        {"$group": {"_id": "$vendor", "paid": {"$sum": "$usd_amount"}}}
    ]).to_list(5000)
    p_map = {p["_id"]: p["paid"] for p in payments_agg if p["_id"]}
    ca_rows = []
    for r in payables_agg:
        if not r["_id"]: continue
        debt = float(r["debt"] or 0)
        paid = float(p_map.get(r["_id"], 0))
        ca_rows.append({
            "name": r["_id"], "debt": debt, "paid": paid,
            "balance": debt - paid, "count": r["count"],
        })
    ca_rows.sort(key=lambda x: x["balance"], reverse=True)
    _write_sheet(wb, "Cari Hesaplar", ca_rows, [
        ("name", "Cari Adı"), ("debt", "Toplam Borç (USD)"),
        ("paid", "Ödenen (USD)"), ("balance", "Bakiye (USD)"),
        ("count", "Kayıt Sayısı"),
    ])

    # 5. Master Data — Tedarikçiler
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(5000)
    _write_sheet(wb, "Tedarikçiler", vendors, [
        ("name", "Ad"), ("country", "Ülke"), ("tax_no", "Vergi No"),
        ("iban", "IBAN"), ("contact", "İletişim"), ("phone", "Telefon"),
        ("email", "Email"), ("notes", "Notlar"),
    ])

    # 6. Gemiler
    ships = await db.ships.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Gemiler", ships, [
        ("name", "Gemi/Birim"), ("manager", "Manager"), ("armator", "Armatör"),
        ("imo", "IMO"), ("flag", "Bayrak"), ("notes", "Notlar"),
    ])

    # 7. Bankalar
    banks = await db.banks.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Bankalar", banks, [
        ("name", "Banka/Kasa Adı"), ("type", "Tip"), ("currency", "Para Birimi"),
        ("iban", "IBAN"), ("balance", "Bakiye"), ("company", "Şirket"),
    ])

    # 8. Şirketler
    companies = await db.companies.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Şirketler", companies, [
        ("name", "Şirket Adı"), ("tax_no", "Vergi No"), ("notes", "Notlar"),
    ])

    # 9. Masraf Türleri
    exp_types = await db.expense_types.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Masraf Türleri", exp_types, [
        ("code", "Kod"), ("name", "Ad"), ("notes", "Notlar"),
    ])

    # 10. Muhasebe Hesap Planı
    acc_codes = await db.accounting_codes.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Hesap Planı", acc_codes, [
        ("code", "Kod"), ("name", "Hesap Adı"),
    ])

    # 11. Döviz Kurları
    fx = await db.currencies.find({}, {"_id": 0}).to_list(50)
    _write_sheet(wb, "Döviz Kurları", fx, [
        ("code", "Kod"), ("name", "Ad"), ("rate_to_tl", "TL Karşılığı"),
        ("forex_selling", "Döviz Satış"), ("banknote_buying", "Efektif Alış"),
        ("banknote_selling", "Efektif Satış"), ("last_updated", "Son Güncelleme"),
    ])

    # 12. Armatörler
    armators = await db.armators.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Armatörler", armators, [("name", "Ad"), ("notes", "Notlar")])

    # 13. Kişi & Alt Şirketler
    people = await db.people.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Kişi & Şirketler", people, [("name", "Ad"), ("notes", "Notlar")])

    # 14. Ülkeler
    countries = await db.countries.find({}, {"_id": 0}).to_list(500)
    _write_sheet(wb, "Ülkeler", countries, [("name", "Ad"), ("code", "Kod")])

    # Stream as response
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"marti-finans-veri-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    await write_audit(user, "export", "excel", meta={"filename": filename})
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
